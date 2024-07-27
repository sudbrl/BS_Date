import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from io import BytesIO
import requests

# Mock function to convert English date to Nepali date
def convert_to_nepali_date(eng_date):
    try:
        # Placeholder logic for conversion
        np_year, np_month, np_day = eng_date.year + 56, eng_date.month, eng_date.day
        return f'{np_year}-{np_month:02d}-{np_day:02d}'
    except Exception:
        return 'Out of range'

# Function to extract year and month from Nepali date
def extract_year_month(np_date):
    if np_date == 'Out of range':
        return None, None
    np_year, np_month, _ = map(int, np_date.split('-'))
    return np_year, np_month

# Function to calculate fiscal year and quarter
def calculate_fy_quarter(np_date):
    if np_date == 'Out of range':
        return 'Invalid Format'
    
    np_year, np_month = extract_year_month(np_date)
    
    if np_year is None or np_month is None:
        return 'Invalid Format'

    if 1 <= np_month <= 3:
        fy_quarter = f"FY{np_year - 1}/{np_year} Q4"
    elif 4 <= np_month <= 6:
        fy_quarter = f"FY{np_year}/{np_year + 1} Q1"
    elif 7 <= np_month <= 9:
        fy_quarter = f"FY{np_year}/{np_year + 1} Q2"
    elif 10 <= np_month <= 12:
        fy_quarter = f"FY{np_year}/{np_year + 1} Q3"
    else:
        fy_quarter = 'Invalid Format'

    return fy_quarter

# Function to map Nepali month number to name
def map_month_to_name(np_date):
    if np_date == 'Out of range':
        return 'Invalid Date'
    
    _, np_month = extract_year_month(np_date)
    
    month_mapping = {
        1: "Baisakh",
        2: "Jestha",
        3: "Ashadh",
        4: "Shrawan",
        5: "Bhadra",
        6: "Ashwin",
        7: "Kartik",
        8: "Mangsir",
        9: "Poush",
        10: "Magh",
        11: "Falgun",
        12: "Chaitra"
    }
    
    return month_mapping.get(np_month, 'Invalid Month')

# Streamlit app
st.set_page_config(page_title="Nepali Date Processing App", page_icon="📅", layout="wide")
st.title("📅 Nepali Date Processing App")

# Custom CSS to make the display more compact
st.markdown(
    """
    <style>
    .css-1v3fvcr { font-size: 10px; } /* Smaller font size */
    .css-1r7g26i { width: 80%; margin: auto; } /* Reduced width */
    .css-1emrehy { font-size: 10px; } /* Smaller font size for other elements */
    .stButton { padding: 5px; font-size: 10px; } /* Reduced padding and font size for buttons */
    </style>
    """,
    unsafe_allow_html=True
)

# Provide a link to download a sample workbook
sample_url = "https://github.com/sudbrl/BS_Date/blob/main/BS_Date.xlsx"
st.markdown(f"Download a sample workbook [here]({sample_url}).")

# File upload
uploaded_file = st.file_uploader("Upload an Excel file", type="xlsx")

# Button to start processing
if uploaded_file:
    if st.button("Begin Processing"):
        with st.spinner("Processing..."):
            try:
                # Read the uploaded Excel file
                df = pd.read_excel(uploaded_file, sheet_name=None)

                df_list = []

                for sheet_name, sheet_df in df.items():
                    # Ensure the date column is in datetime format without time
                    sheet_df.iloc[:, 0] = pd.to_datetime(sheet_df.iloc[:, 0].apply(lambda x: str(x).split(' ')[0]), format='%Y-%m-%d', errors='coerce')
                    sheet_df['Nepali Date'] = sheet_df.iloc[:, 0].apply(lambda x: convert_to_nepali_date(x) if pd.notnull(x) else 'Out of range')
                    sheet_df['Fiscal Year and Quarter'] = sheet_df['Nepali Date'].apply(calculate_fy_quarter)
                    sheet_df['BS_Month'] = sheet_df['Nepali Date'].apply(map_month_to_name)
                    df_list.append(sheet_df)

                df = pd.concat(df_list, ignore_index=True)

                # Save to a BytesIO object
                output = BytesIO()
                df.to_excel(output, index=False, engine='openpyxl')

                # Formatting the date in the output Excel file
                output.seek(0)
                wb = load_workbook(output)
                ws = wb.active

                # Define a date style
                date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')

                # Apply the date style to the 'Nepali Date' column
                for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.style = date_style

                # Save the formatted Excel file
                formatted_output = BytesIO()
                wb.save(formatted_output)
                formatted_output.seek(0)

                st.success("File processed successfully!")

                # Provide a download link
                st.download_button(
                    label="Download Processed File",
                    data=formatted_output,
                    file_name="BS_Date_converted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"Error processing file: {e}")
